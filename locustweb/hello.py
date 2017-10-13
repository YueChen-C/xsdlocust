#coding=utf8
import json
import os

import openpyxl
from flask import Flask, render_template,request, flash
from flask_script import Manager
from flask_bootstrap import Bootstrap
from flask_moment import Moment
global null
null=''


app = Flask(__name__)
app.debug=True
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SECRET_KEY'] = os.urandom(24)

manager = Manager(app)
bootstrap = Bootstrap(app)
moment = Moment(app)
def datalist():
    readpath = os.getcwd() + '/interface.xlsx'
    inwb = openpyxl.load_workbook(readpath)
    work_name= inwb.get_sheet_by_name('Sheet1')

    list=[]
    for row in work_name.rows:
        key=[]
        for cell in row:
            key.append(cell.value)
        list.append(key)
    return list
def convert(input):
    if isinstance(input, dict):
        return {convert(key): convert(value) for key, value in input.iteritems()}
    elif isinstance(input, list):
        return [convert(element) for element in input]
    elif isinstance(input, unicode):
        return input.encode('utf-8')
    else:
        return input



@app.route('/index',methods=['GET', 'POST'])
def index():
    hobby = request.form.get('keydata')
    readpath = os.getcwd() + '/interface.xlsx'
    inwb = openpyxl.load_workbook(readpath)
    work_name= inwb.get_sheet_by_name('Sheet1')
    hobby=json.loads(json.dumps(hobby,ensure_ascii=False))

    if hobby:
        for key in hobby:
            flash(key)
    # for i in range(1,len(work_name.rows)+1):
    #     if str(i) in hobby:
    #         work_name.cell(row = i,column = 5).value=1
    #     else:
    #         work_name.cell(row = i,column = 5).value=0
    # inwb.save(readpath)
    data=datalist()

    return render_template('index.html',datalist=data)


# @app.route('/index',methods=['GET', 'POST'])
# def index():
#
#     data=datalist()
#     hobby = request.form.getlist('keydata')
#     if request.method!='POST':
#         pass
#     elif hobby==[] :
#         flash(hobby)
#         # flash(u'提示：还未选择测试接口，请选择测试接口')
#     else:
#         readpath = os.getcwd() + '/interface.xlsx'
#         inwb = openpyxl.load_workbook(readpath)
#         work_name= inwb.get_sheet_by_name('Sheet1')
#         for i in range(1,len(work_name.rows)+1):
#             if str(i) in hobby:
#                 work_name.cell(row = i,column = 5).value=1
#             else:
#                 work_name.cell(row = i,column = 5).value=0
#         inwb.save(readpath)
#         data=datalist()
#         return render_template('index.html',datalist=data)
#     return render_template('index.html',datalist=data)











if __name__ == '__main__':
    manager.run()