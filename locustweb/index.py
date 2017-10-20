# coding=utf8
import ConfigParser
import json
import os
import random
from time import sleep
import MySQLdb
import openpyxl
import psutil
import redis
import requests
import subprocess
from flask import Flask, render_template, request, flash
from flask_bootstrap import Bootstrap
from flask_moment import Moment
from flask_script import Manager
import common

global null
null = ''

app = Flask(__name__)
app.debug = True
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SECRET_KEY'] = os.urandom(24)
manager = Manager(app)
bootstrap = Bootstrap(app)
moment = Moment(app)

cf = ConfigParser.ConfigParser()
filecfg = os.path.dirname(os.path.realpath(__file__)) + '/config.cfg'
filexlsx = os.path.dirname(os.path.realpath(__file__)) + '/interface.xlsx'


def datalist():
    readpath = filexlsx
    inwb = openpyxl.load_workbook(readpath)
    work_name = inwb.get_sheet_by_name('Sheet1')
    list = []
    for row in range(1, work_name.max_row + 1):
        k1 = work_name.cell(row=row, column=1).value
        k2 = work_name.cell(row=row, column=2).value
        postdata = eval(work_name.cell(row=row, column=3).value)
        k3 = postdata['b']
        k4 = postdata['h']['version']
        k5 = work_name.cell(row=row, column=5).value
        key = [k1, k2, k3, k4, k5]
        list.append(key)
    return list


@app.route('/index', methods=['GET', 'POST'])
def index():
    louststatus = False
    status = False
    hobby = request.form.get('keydata')
    test = request.form.get('test')
    runserver = request.form.get('runlocust')
    killserver = request.form.get('killlocust')
    ticket = request.form.get('ticket')
    locustcfg = common.cfg(filecfg, 'locust')
    if ticket:
        try:
            url = "https://test3.txdsd.com/platform-rest/service.jws"
            # 清理ticket
            rediscfg = common.cfg(filecfg, 'redis')
            pool = redis.Redis(host=rediscfg.query('host'), port=rediscfg.query('port'), db=rediscfg.query('db'),
                               password=rediscfg.query('password'))
            keys = pool.keys('h-member-session*') + pool.keys('v-session-appToken*') + pool.keys('v-session-ticket*')
            if keys:
                pool.delete(*keys)

            # 更新登录信息
            db = common.cfg(filecfg, 'mysql')
            conn = MySQLdb.connect(host=db.query('host'), user=db.query('user'), passwd=db.query('passwd'),
                                   db=db.query('db'), charset=db.query('charset'))
            cursors = conn.cursor()
            ticketnum = common.cfg(filecfg, 'locust').query('ticketnum')
            cursors.execute("SELECT mobile_phone,login_password FROM member ORDER BY member_id  DESC  LIMIT {0}".format(ticketnum))
            dbdata = cursors.fetchall()
            cursors.close()
            conn.close()
            for key in dbdata:
                login = common.logindata
                login['b']['loginName'] = key[0]
                login['b']['loginPassword'] = key[1]
                requests.post(url=url, json=login)
            flash(u'提示：{}条ticket登录信息更新成功'.format(ticketnum), 'warning')
        except Exception as p:
            flash(u'警告：更新失败{0}'.format(p), 'danger')

    # 检查locust服务是否启动
    for proc in psutil.process_iter(attrs=['pid', 'name']):
        if proc.info['name'] == 'locust.exe':
            louststatus = True

    #启动locust服务
    if runserver:
        try:
            data = datalist()
            configdata = locustcfg.cfgdict()
            for proc in psutil.process_iter(attrs=['pid', 'name']):
                if proc.info['name'] == 'locust.exe':
                    louststatus = True
                    flash(u'警告：locust服务已经启动，请停止后再启动！', 'danger')
                    return render_template('index.html', datalist=data, configdata=configdata, louststatus=louststatus)
            subprocess.Popen("locust --master -f {0}/test.py ".format(os.path.dirname(os.path.realpath(__file__))))
            subprocess.Popen("python {0}/openslave.py ".format(os.path.dirname(os.path.realpath(__file__))))
            sleep(3)
            louststatus = True
            status = True
            flash(u'成功：locust服务已经启动', 'success')
            return render_template('index.html', datalist=data, configdata=configdata, louststatus=louststatus,
                                   status=status)
        except Exception as E:
            flash(u'警告：locust启动失败 {0}'.format(E))

    #停止locust服务
    if killserver:
        for proc in psutil.process_iter(attrs=['pid', 'name']):
            if proc.info['name'] == 'locust.exe':
                p = psutil.Process(proc.info['pid'])
                psutil.Process(p.children()[0].pid).terminate()
                louststatus = False
        flash(u'提示：locust服务已经关闭', 'warning')

    #所选接口测试
    if test:
        try:
            readpath = filexlsx
            inwb = openpyxl.load_workbook(readpath)
            work_name = inwb.get_sheet_by_name('Sheet1')
            faillist = {}
            ticketlist = common.exportTicket(num=10)
            if ticketlist:
                ticketdata = random.choice(ticketlist)
                for row in range(1, work_name.max_row + 1):
                    if work_name.cell(row=row, column=5).value == 1:
                        name = work_name.cell(row=row, column=2).value
                        postdata = eval(work_name.cell(row=row, column=3).value)
                        postdata['h']['appToken'] = ticketdata[1]
                        postdata['h']['ticket'] = ticketdata[0]
                        data = requests.post('https://test3.txdsd.com/platform-rest/service.jws', json=postdata).json()
                        if data['h']['code'] == '0':
                            pass
                        else:
                            faillist[name] = data['h']['message']
            else:
                configdata = locustcfg.cfgdict()
                data = datalist()
                flash(u'警告：当前没有正在登录的用户，获取ticket失败', 'danger')
                return render_template('index.html', datalist=data, configdata=configdata, louststatus=louststatus)

            if faillist:
                flash(u'警告：接口测试失败：{0}'.format(json.dumps(faillist, ensure_ascii=False, indent=2)), 'danger')
            else:
                flash(u'成功：所选接口测试通过', 'success')
        except Exception as a:
            flash(u'警告：{0}'.format(a), 'danger')


    #保存修改数据
    if hobby:
        try:
            readpath = filexlsx
            inwb = openpyxl.load_workbook(readpath)
            work_name = inwb.get_sheet_by_name('Sheet1')
            hobby = hobby.encode('unicode-escape').decode('string_escape')
            hobby = json.loads(hobby)
            for key, value in hobby.items():
                postdata = eval(work_name.cell(row=int(key), column=3).value)
                postdata['b'] = eval(value[0].encode('unicode-escape'))
                postdata['h']['version'] = value[1].encode('unicode-escape')
                work_name.cell(row=int(key), column=3).value = str(postdata)
                work_name.cell(row=int(key), column=5).value = value[2]
            inwb.save(readpath)
            flash(u'成功：保存成功', 'success')
        except Exception, E:
            flash(u'警告：保存失败{0}'.format(E), 'danger')
    data = datalist()
    configdata = locustcfg.cfgdict()
    return render_template('index.html', datalist=data, configdata=configdata, louststatus=louststatus)


if __name__ == '__main__':
    manager.run()
