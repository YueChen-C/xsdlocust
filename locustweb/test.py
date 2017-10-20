# coding=utf-8
import os
import random

import openpyxl
from locust import HttpLocust, TaskSet, task
import common


def postdata():
    readpath = os.path.dirname(os.path.realpath(__file__)) + '/interface.xlsx'
    inwb = openpyxl.load_workbook(readpath)
    work_name = inwb.get_sheet_by_name('Sheet1')
    list = []
    for row in range(1, work_name.max_row + 1):
        if work_name.cell(row=row, column=5).value == 1:
            k2 = work_name.cell(row=row, column=2).value
            jsondata = eval(work_name.cell(row=row, column=3).value)
            key = [k2, jsondata]
            list.append(key)
    return list


class WebsiteTasks(TaskSet):
    def on_start(self):
        self.ticketlist = random.choice(self.locust.ticketlist)

    @task
    def first(self):
        data = random.choice(self.locust.postdata)
        data[1]['h']['appToken'] = self.ticketlist[1]
        data[1]['h']['ticket'] = self.ticketlist[0]
        with self.client.post("/platform-rest/service.jws", json=data[1], name=data[0],
                              catch_response=True) as response:
            if response.status_code == 200:
                response.success()
            else:
                response.failure(u'请求失败')


class WebsiteUser(HttpLocust):
    file = os.path.dirname(os.path.realpath(__file__)) + '/config.cfg'
    config = common.cfg(file, 'locust')
    task_set = WebsiteTasks
    host = config.query('host')
    ticketlist = common.exportTicket(num=int(config.query('ticketnum')))
    postdata = postdata()
    min_waittime = int(config.query('min_waittime'))
    max_waittime = int(config.query('max_waittime'))
